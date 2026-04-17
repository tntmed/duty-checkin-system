"""Dashboard and timeline router."""
from datetime import date, datetime
from io import BytesIO
from typing import List, Optional

from fastapi import APIRouter, Depends, HTTPException, Query, status
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from sqlalchemy import func, case
from sqlalchemy.orm import Session

from ..db import get_db
from .. import models, schemas
from .auth import get_current_user

router = APIRouter()


# ============================================================
# Permission helper
# ============================================================

def require_role(*required_roles: str):
    """
    FastAPI dependency factory.
    Raises 403 if the current user has none of the required role names.
    Pass no roles to allow any authenticated user.
    """
    def _check(current_user: models.User = Depends(get_current_user)) -> models.User:
        if not required_roles:
            return current_user
        user_role_names = {r.name for r in current_user.roles}
        if not user_role_names.intersection(required_roles):
            raise HTTPException(
                status_code=status.HTTP_403_FORBIDDEN,
                detail=f"Access denied. Required role: {' or '.join(required_roles)}",
            )
        return current_user
    return _check


# ============================================================
# Dashboard: GET /dashboard/summary
# ============================================================

@router.get("/dashboard/summary", tags=["Dashboard"])
def get_dashboard_summary(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """KPI summary for today: total duties, open incidents, checklist issues, pending checkout."""
    today = date.today()

    total_duties = (
        db.query(func.count(models.Duty.id))
        .filter(models.Duty.duty_date == today)
        .scalar() or 0
    )

    open_incidents = (
        db.query(func.count(models.Incident.id))
        .join(models.Duty, models.Incident.duty_id == models.Duty.id)
        .filter(
            models.Duty.duty_date == today,
            models.Incident.status.in_(["OPEN", "IN_PROGRESS"]),
        )
        .scalar() or 0
    )

    checklist_issues = (
        db.query(func.count(models.ChecklistLog.id))
        .join(models.Duty, models.ChecklistLog.duty_id == models.Duty.id)
        .filter(
            models.Duty.duty_date == today,
            models.ChecklistLog.status == "NOT_OK",
        )
        .scalar() or 0
    )

    pending_checkout = (
        db.query(func.count(models.Duty.id))
        .filter(
            models.Duty.duty_date == today,
            models.Duty.checkout_time.is_(None),
        )
        .scalar() or 0
    )

    return {
        "total_duties": total_duties,
        "open_incidents": open_incidents,
        "checklist_issues": checklist_issues,
        "pending_checkout": pending_checkout,
    }


# ============================================================
# Dashboard: GET /dashboard/duties
# ============================================================

@router.get("/dashboard/duties", tags=["Dashboard"])
def get_dashboard_duties(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
    duty_date: Optional[date] = Query(None, description="Filter by date (YYYY-MM-DD). Defaults to today."),
    role_id: Optional[int] = Query(None, description="Filter by role ID"),
    shift_id: Optional[int] = Query(None, description="Filter by shift ID"),
    attendance_status: Optional[str] = Query(None, description="Filter by attendance status"),
    has_issues: Optional[bool] = Query(None, description="If true, return only duties with NOT OK checklist items"),
    has_incidents: Optional[bool] = Query(None, description="If true, return only duties with incidents"),
    pending_checkout: Optional[bool] = Query(None, description="If true, return only duties not yet checked out"),
):
    """
    List duties with issue_count and incident_count.
    Supports filters: duty_date, role_id, shift_id, attendance_status,
    has_issues, has_incidents, pending_checkout.
    """
    today = duty_date or date.today()

    # Fetch duties with joins — count issues/incidents in Python to avoid Oracle SQL edge cases
    query = (
        db.query(models.Duty)
        .join(models.User, models.Duty.user_id == models.User.id)
        .join(models.Role, models.Duty.role_id == models.Role.id)
        .join(models.DutyShift, models.Duty.shift_id == models.DutyShift.id)
        .filter(models.Duty.duty_date == today)
    )

    if role_id is not None:
        query = query.filter(models.Duty.role_id == role_id)
    if shift_id is not None:
        query = query.filter(models.Duty.shift_id == shift_id)
    if attendance_status:
        query = query.filter(models.Duty.attendance_status == attendance_status)
    if pending_checkout is True:
        query = query.filter(models.Duty.checkout_time.is_(None))
    if pending_checkout is False:
        query = query.filter(models.Duty.checkout_time.isnot(None))

    duties = query.order_by(models.Duty.id.desc()).all()

    result = []
    for duty in duties:
        issue_count = sum(1 for log in duty.checklist_logs if log.status == "NOT_OK")
        incident_count = len(duty.incidents)
        result.append({
            "duty_id": duty.id,
            "user_id": duty.user_id,
            "duty_date": str(duty.duty_date),
            "employee_code": duty.user.employee_code,
            "full_name": duty.user.full_name,
            "role_name": duty.role.name,
            "shift_name": duty.shift.name,
            "checkin_time": duty.checkin_time.isoformat() if duty.checkin_time else None,
            "checkout_time": duty.checkout_time.isoformat() if duty.checkout_time else None,
            "attendance_status": duty.attendance_status,
            "attendance_confirmed": bool(duty.attendance_confirmed),
            "attendance_confirmed_at": duty.attendance_confirmed_at.isoformat() if duty.attendance_confirmed_at else None,
            "issue_count": issue_count,
            "incident_count": incident_count,
        })

    if has_issues is True:
        result = [r for r in result if r["issue_count"] > 0]
    if has_incidents is True:
        result = [r for r in result if r["incident_count"] > 0]

    return result




# ============================================================
# Timeline: GET /duties/{duty_id}/timeline
# ============================================================

@router.get("/duties/{duty_id}/timeline", response_model=schemas.TimelineResponse, tags=["Timeline"])
def get_duty_timeline(
    duty_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """
    Build a chronological activity timeline for a duty.
    Aggregates: check-in, checklist status changes, incidents, checkout.
    """
    duty = db.query(models.Duty).filter(models.Duty.id == duty_id).first()
    if not duty:
        from fastapi import HTTPException, status
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Duty not found")

    events: List[schemas.TimelineEvent] = []

    # 1. Check-in event
    if duty.checkin_time:
        events.append(schemas.TimelineEvent(
            timestamp=duty.checkin_time,
            event_type="CHECKIN",
            label="Checked in",
            detail=f"Role: {duty.role.name if duty.role else '—'} | Shift: {duty.shift.name if duty.shift else '—'}",
            severity="info",
        ))

    # 2. Checklist events (only items that have been acted on: not still in initial NA with no checked_at)
    checklist_logs = (
        db.query(models.ChecklistLog)
        .filter(models.ChecklistLog.duty_id == duty_id)
        .all()
    )
    for log in checklist_logs:
        # Use checked_at if set, else skip (never touched)
        ts = log.checked_at
        if ts is None:
            continue
        item_text = log.template.item_text if log.template else f"Item #{log.template_id}"
        order = log.template.item_order if log.template else "?"

        if log.status == "OK":
            severity = "success"
            label = f"Checklist #{order} — OK"
        elif log.status == "NOT_OK":
            severity = "danger"
            label = f"Checklist #{order} — NOT OK"
        else:
            severity = "info"
            label = f"Checklist #{order} — N/A"

        events.append(schemas.TimelineEvent(
            timestamp=ts,
            event_type="CHECKLIST",
            label=label,
            detail=f"{item_text}" + (f" | Remark: {log.remark}" if log.remark else ""),
            severity=severity,
        ))

    # 3. Incident events
    incidents = (
        db.query(models.Incident)
        .filter(models.Incident.duty_id == duty_id)
        .all()
    )
    for inc in incidents:
        impact_label = f"[{inc.impact}]" if inc.impact != "NONE" else ""
        severity = {
            "NONE": "info",
            "MINOR": "warning",
            "MAJOR": "warning",
            "CRITICAL": "danger",
        }.get(inc.impact, "info")

        events.append(schemas.TimelineEvent(
            timestamp=inc.reported_at,
            event_type="INCIDENT",
            label=f"Incident reported — {inc.incident_type} {impact_label}".strip(),
            detail=str(inc.detail)[:120] + ("..." if len(str(inc.detail)) > 120 else ""),
            severity=severity,
        ))

        if inc.resolved_at:
            events.append(schemas.TimelineEvent(
                timestamp=inc.resolved_at,
                event_type="INCIDENT_RESOLVED",
                label=f"Incident resolved — {inc.incident_type}",
                severity="success",
            ))

    # 4. Checkout event
    if duty.checkout_time:
        events.append(schemas.TimelineEvent(
            timestamp=duty.checkout_time,
            event_type="CHECKOUT",
            label="Checked out",
            severity="info",
        ))

    # Sort chronologically
    events.sort(key=lambda e: e.timestamp)

    return schemas.TimelineResponse(duty_id=duty_id, events=events)


# ============================================================
# Dashboard: GET /dashboard/export  (CSV download)
# ============================================================

@router.get("/dashboard/export", tags=["Dashboard"])
def export_dashboard_csv(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
    duty_date: Optional[date] = Query(None, description="Single date filter (legacy). Ignored when date_from/date_to are set."),
    date_from: Optional[date] = Query(None, description="Range start date (YYYY-MM-DD)"),
    date_to: Optional[date] = Query(None, description="Range end date (YYYY-MM-DD)"),
    role_id: Optional[int] = Query(None),
    shift_id: Optional[int] = Query(None),
    attendance_status: Optional[str] = Query(None),
    has_issues: Optional[bool] = Query(None),
    has_incidents: Optional[bool] = Query(None),
    pending_checkout: Optional[bool] = Query(None),
):
    """Export duties as Excel. Supports single date or date range (date_from / date_to)."""
    today = date.today()

    # Resolve effective date range
    if date_from or date_to:
        eff_from = date_from or (date_to or today)
        eff_to = date_to or (date_from or today)
    else:
        eff_from = eff_to = duty_date or today

    if eff_from > eff_to:
        eff_from, eff_to = eff_to, eff_from

    query = (
        db.query(models.Duty)
        .join(models.User, models.Duty.user_id == models.User.id)
        .join(models.Role, models.Duty.role_id == models.Role.id)
        .join(models.DutyShift, models.Duty.shift_id == models.DutyShift.id)
        .filter(models.Duty.duty_date >= eff_from, models.Duty.duty_date <= eff_to)
    )
    if role_id is not None:
        query = query.filter(models.Duty.role_id == role_id)
    if shift_id is not None:
        query = query.filter(models.Duty.shift_id == shift_id)
    if attendance_status:
        query = query.filter(models.Duty.attendance_status == attendance_status)
    if pending_checkout is True:
        query = query.filter(models.Duty.checkout_time.is_(None))
    if pending_checkout is False:
        query = query.filter(models.Duty.checkout_time.isnot(None))

    duties = query.order_by(models.Duty.id.desc()).all()

    rows = []
    for duty in duties:
        issue_count = sum(1 for log in duty.checklist_logs if log.status == "NOT_OK")
        incident_count = len(duty.incidents)
        rows.append({
            "duty_id": duty.id,
            "duty_date": str(duty.duty_date),
            "employee_code": duty.user.employee_code,
            "full_name": duty.user.full_name,
            "role_name": duty.role.name,
            "shift_name": duty.shift.name,
            "checkin_time": duty.checkin_time.strftime("%d/%m/%Y %H:%M") if duty.checkin_time else "",
            "checkout_time": duty.checkout_time.strftime("%d/%m/%Y %H:%M") if duty.checkout_time else "",
            "attendance_status": duty.attendance_status or "",
            "issue_count": issue_count,
            "incident_count": incident_count,
        })

    if has_issues is True:
        rows = [r for r in rows if r["issue_count"] > 0]
    if has_incidents is True:
        rows = [r for r in rows if r["incident_count"] > 0]

    # Build Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "รายงานการลงเวร"

    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        ("ลำดับ", 8), ("วันที่", 14), ("รหัส", 12), ("ชื่อ-นามสกุล", 28),
        ("ตำแหน่ง/Role", 20), ("เวร/Shift", 22), ("เวลา Check-in", 18),
        ("เวลา Check-out", 18), ("สถานะ", 12), ("ปัญหา Checklist", 18), ("Incidents", 12),
    ]
    for col, (h, w) in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
        ws.column_dimensions[cell.column_letter].width = w

    ws.row_dimensions[1].height = 22

    status_th = {"PRESENT": "มาปกติ", "LATE": "มาสาย", "ABSENT": "ขาด", "EXCUSED": "ลา"}

    for i, r in enumerate(rows, 1):
        row_data = [
            i, r["duty_date"], r["employee_code"], r["full_name"],
            r["role_name"], r["shift_name"], r["checkin_time"], r["checkout_time"],
            status_th.get(r["attendance_status"], r["attendance_status"]),
            r["issue_count"], r["incident_count"],
        ]
        bg = "FFFFFF" if i % 2 == 0 else "F0F4FF"
        fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=i + 1, column=col, value=val)
            cell.border = border
            cell.fill = fill
            cell.alignment = Alignment(vertical="center")

    # Sheet 2: Incidents
    ws2 = wb.create_sheet("Incidents")
    inc_headers = [
        ("ลำดับ", 8), ("วันที่เวร", 14), ("ชื่อผู้เวร", 28), ("ตำแหน่ง/Role", 20),
        ("ประเภท Incident", 20), ("ความรุนแรง", 16), ("รายละเอียด", 50),
        ("สถานะ", 14), ("เวลารายงาน", 20), ("เวลาแก้ไข", 20), ("หมายเหตุการแก้ไข", 40),
    ]
    for col, (h, w) in enumerate(inc_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border
        ws2.column_dimensions[cell.column_letter].width = w
    ws2.row_dimensions[1].height = 22

    inc_row = 2
    impact_th = {"NONE": "ไม่มี", "MINOR": "น้อย", "MAJOR": "ปานกลาง", "CRITICAL": "วิกฤต"}
    inc_status_th = {"OPEN": "เปิด", "IN_PROGRESS": "กำลังดำเนินการ", "RESOLVED": "แก้ไขแล้ว", "CLOSED": "ปิด"}
    for duty in duties:
        for inc in duty.incidents:
            bg = "FFFFFF" if inc_row % 2 == 0 else "FFF8F0"
            fill2 = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
            row_data = [
                inc_row - 1,
                str(duty.duty_date),
                duty.user.full_name,
                duty.role.name,
                inc.incident_type,
                impact_th.get(inc.impact, inc.impact),
                str(inc.detail),
                inc_status_th.get(inc.status, inc.status),
                inc.reported_at.strftime("%d/%m/%Y %H:%M") if inc.reported_at else "",
                inc.resolved_at.strftime("%d/%m/%Y %H:%M") if inc.resolved_at else "",
                inc.resolution_note or "",
            ]
            for col, val in enumerate(row_data, 1):
                cell = ws2.cell(row=inc_row, column=col, value=val)
                cell.border = border
                cell.fill = fill2
                cell.alignment = Alignment(vertical="center", wrap_text=True)
            inc_row += 1

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"duties_{eff_from}_to_{eff_to}.xlsx" if eff_from != eff_to else f"duties_{eff_from}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ============================================================
# Dashboard: GET /dashboard/today
# ============================================================

@router.get("/dashboard/today", response_model=schemas.DashboardResponse, tags=["Dashboard"])
def get_dashboard_today(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    """
    Manager dashboard for today: NOT OK checklist items, open incidents,
    late check-ins, and list of all active users with their check-in status.
    """
    today = date.today()

    # All duties today
    duties_today = (
        db.query(models.Duty)
        .filter(models.Duty.duty_date == today)
        .all()
    )

    # NOT OK checklist items from today's duties
    not_ok_items: List[schemas.DashboardChecklistIssue] = []
    for duty in duties_today:
        for log in duty.checklist_logs:
            if log.status == "NOT_OK":
                not_ok_items.append(schemas.DashboardChecklistIssue(
                    duty_id=duty.id,
                    user_name=duty.user.full_name if duty.user else "—",
                    role_name=duty.role.name if duty.role else "—",
                    item_text=log.template.item_text if log.template else f"Item #{log.template_id}",
                    remark=log.remark,
                ))

    # Open incidents from today's duties
    open_incidents: List[schemas.DashboardIncident] = []
    for duty in duties_today:
        for inc in duty.incidents:
            if inc.status in ("OPEN", "IN_PROGRESS"):
                open_incidents.append(schemas.DashboardIncident(
                    incident_id=inc.id,
                    duty_id=duty.id,
                    user_name=duty.user.full_name if duty.user else "—",
                    incident_type=inc.incident_type,
                    impact=inc.impact,
                    detail=str(inc.detail)[:200],
                    reported_at=inc.reported_at,
                ))

    # Checked-in user IDs today
    checked_in_user_ids = {d.user_id for d in duties_today}

    # All active users
    all_users = db.query(models.User).filter(models.User.is_active == 1).all()

    user_statuses: List[schemas.DashboardUserStatus] = []
    for u in all_users:
        # Find their latest duty today (if any)
        user_duty = next(
            (d for d in sorted(duties_today, key=lambda x: x.created_at or datetime.min, reverse=True)
             if d.user_id == u.id),
            None,
        )
        user_statuses.append(schemas.DashboardUserStatus(
            user_id=u.id,
            employee_code=u.employee_code,
            full_name=u.full_name,
            checked_in=u.id in checked_in_user_ids,
            attendance_status=user_duty.attendance_status if user_duty else None,
            role_name=user_duty.role.name if user_duty and user_duty.role else None,
            checkin_time=user_duty.checkin_time if user_duty else None,
        ))

    # Sort: not checked-in first, then late, then present
    def sort_key(us: schemas.DashboardUserStatus):
        if not us.checked_in:
            return 0
        if us.attendance_status == "LATE":
            return 1
        return 2

    user_statuses.sort(key=sort_key)

    late_count = sum(1 for d in duties_today if d.attendance_status == "LATE")

    return schemas.DashboardResponse(
        date=today,
        total_duties=len(duties_today),
        not_ok_count=len(not_ok_items),
        open_incident_count=len(open_incidents),
        late_count=late_count,
        not_checked_in_count=len(all_users) - len(checked_in_user_ids),
        not_ok_items=not_ok_items,
        open_incidents=open_incidents,
        user_statuses=user_statuses,
    )
