"""User management router."""
from typing import List
from io import BytesIO
import secrets
import string

from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from ..db import get_db
from .. import crud, schemas, models
from .auth import get_current_user

router = APIRouter()


def require_admin(current_user: models.User = Depends(get_current_user)) -> models.User:
    if not int(current_user.is_admin):
        raise HTTPException(status_code=status.HTTP_403_FORBIDDEN, detail="Admin only")
    return current_user


def _build_user_response(user: models.User, roles: list) -> schemas.UserResponse:
    """Helper to build a UserResponse with roles."""
    return schemas.UserResponse(
        id=user.id,
        employee_code=user.employee_code,
        full_name=user.full_name,
        email=user.email,
        is_active=user.is_active,
        created_at=user.created_at,
        updated_at=user.updated_at,
        roles=[
            schemas.RoleResponse(
                id=r.id,
                name=r.name,
                description=r.description,
                is_active=r.is_active,
                created_at=r.created_at,
            )
            for r in roles
        ],
    )


@router.get("/me", response_model=schemas.UserResponse)
def get_my_profile(
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Return the authenticated user's own profile."""
    roles = crud.get_user_roles(db, current_user.id)
    return _build_user_response(current_user, roles)


@router.get("/me/roles", response_model=List[schemas.RoleResponse])
def get_my_roles(
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Return the authenticated user's assigned roles."""
    roles = crud.get_user_roles(db, current_user.id)
    return [
        schemas.RoleResponse(
            id=r.id,
            name=r.name,
            description=r.description,
            is_active=r.is_active,
            created_at=r.created_at,
        )
        for r in roles
    ]


@router.get("/assignable", response_model=List[schemas.UserResponse])
def list_assignable_users(
    current_user: models.User = Depends(get_current_user),
    db: Session = Depends(get_db),
):
    """Return active users for assignment dropdowns (no admin required)."""
    users = crud.get_users(db, skip=0, limit=500)
    result = []
    for user in users:
        roles = crud.get_user_roles(db, user.id)
        result.append(_build_user_response(user, roles))
    return result


@router.get("/", response_model=List[schemas.UserResponse])
def list_users(
    skip: int = 0,
    limit: int = 100,
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """List all users (admin endpoint)."""
    users = crud.get_users(db, skip=skip, limit=limit)
    result = []
    for user in users:
        roles = crud.get_user_roles(db, user.id)
        result.append(_build_user_response(user, roles))
    return result


@router.post("/", response_model=schemas.UserResponse, status_code=status.HTTP_201_CREATED)
def create_user(
    user_in: schemas.UserCreate,
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Create a new user."""
    existing = crud.get_user_by_employee_code(db, employee_code=user_in.employee_code)
    if existing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Employee code '{user_in.employee_code}' is already registered",
        )
    new_user = crud.create_user(db, user=user_in)
    return _build_user_response(new_user, [])


@router.get("/import/template")
def download_import_template(
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Download an Excel template for bulk user import."""
    roles = db.query(models.Role).filter(models.Role.is_active == 1).all()
    role_names = ", ".join(r.name for r in roles)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Users"

    # Header row
    headers = ["employee_code *", "full_name *", "email", "password *", "roles (คั่นด้วย ,)"]
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Example rows
    examples = [
        ("EMP003", "สมชาย ใจดี",    "somchai@example.com",  "password123", "สิบเวร, ช่าง"),
        ("EMP004", "สมหญิง รักไทย", "somying@example.com",  "password123", "ผู้ช่วยสิบเวร"),
        ("EMP005", "วิชัย เก่งมาก",  "wichai@example.com",   "password123", "โปรแกรมเมอร์"),
    ]
    for row_data in examples:
        ws.append(row_data)

    # Column widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 40

    # Note sheet
    ws2 = wb.create_sheet("หมายเหตุ")
    ws2["A1"] = "ชื่อ Role ที่ใช้ได้:"
    ws2["A1"].font = Font(bold=True)
    for i, r in enumerate(roles, 2):
        ws2[f"A{i}"] = r.name
    ws2["A" + str(len(roles) + 3)] = f"ตัวอย่าง: {role_names}"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=users_template.xlsx"},
    )


@router.post("/import/excel")
def import_users_from_excel(
    file: UploadFile = File(...),
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Import users from an uploaded Excel file."""
    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="กรุณาอัปโหลดไฟล์ Excel (.xlsx)")

    content = file.file.read()
    wb = openpyxl.load_workbook(BytesIO(content))
    ws = wb.active

    all_roles = {r.name: r for r in db.query(models.Role).all()}

    created, updated, errors = [], [], []

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue

        employee_code = str(row[0]).strip() if row[0] else None
        full_name     = str(row[1]).strip() if row[1] else None
        email         = str(row[2]).strip() if row[2] else None
        password      = str(row[3]).strip() if row[3] else None
        roles_raw     = str(row[4]).strip() if row[4] else ""

        if not employee_code or not full_name:
            errors.append(f"แถว {row_num}: employee_code และ full_name จำเป็นต้องกรอก")
            continue

        existing = crud.get_user_by_employee_code(db, employee_code)

        if existing:
            # อัปเดต full_name และ email
            existing.full_name = full_name
            if email and email.lower() != "none":
                existing.email = email
            # อัปเดต password ถ้ากรอกมา
            if password and password.lower() != "none":
                existing.hashed_password = crud.get_password_hash(password)
            db.flush()
            target_user = existing
            updated.append(employee_code)
        else:
            target_user = crud.create_user(
                db,
                schemas.UserCreate(
                    employee_code=employee_code,
                    full_name=full_name,
                    email=email if email and email.lower() != "none" else None,
                    password=password or "password123",
                ),
            )
            created.append(employee_code)

        # อัปเดต roles — ลบเดิมทั้งหมดแล้วใส่ใหม่ตาม Excel
        if roles_raw:
            db.query(models.UserRole).filter(models.UserRole.user_id == target_user.id).delete()
            db.flush()
            for role_name in [r.strip() for r in roles_raw.split(",") if r.strip()]:
                if role_name in all_roles:
                    db.add(models.UserRole(user_id=target_user.id, role_id=all_roles[role_name].id))
                else:
                    errors.append(f"แถว {row_num}: ไม่พบ role '{role_name}'")

    db.commit()

    return {
        "created": len(created),
        "updated": len(updated),
        "errors": errors,
        "created_codes": created,
        "updated_codes": updated,
    }


@router.post("/{user_id}/reset-password")
def admin_reset_password(
    user_id: int,
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Admin resets a user's password to a random temporary password."""
    user = crud.get_user(db, user_id)
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")

    # สร้าง password ชั่วคราว เช่น Pcm@4x9z
    chars = string.ascii_letters + string.digits
    temp_pw = "Pcm@" + "".join(secrets.choice(chars) for _ in range(6))

    user.hashed_password = crud.get_password_hash(temp_pw)
    db.commit()
    return {"temp_password": temp_pw, "message": f"รีเซ็ต password ของ {user.full_name} สำเร็จ"}


@router.patch("/{user_id}/toggle-active")
def toggle_user_active(
    user_id: int,
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Toggle a user's active status."""
    user = crud.get_user(db, user_id)
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")
    if user.id == current_user.id:
        raise HTTPException(status_code=400, detail="ไม่สามารถปิดบัญชีตัวเองได้")
    user.is_active = 0 if int(user.is_active) == 1 else 1
    db.commit()
    return {"is_active": int(user.is_active)}


@router.post("/{user_id}/roles/{role_id}", response_model=schemas.UserResponse)
def assign_role(
    user_id: int,
    role_id: int,
    current_user: models.User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    """Assign a role to a user."""
    user = crud.get_user(db, user_id)
    if not user:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="User not found")

    role = crud.get_role(db, role_id)
    if not role:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Role not found")

    crud.assign_role_to_user(db, user_id=user_id, role_id=role_id)
    roles = crud.get_user_roles(db, user_id)
    return _build_user_response(user, roles)
